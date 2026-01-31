from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.views.decorators.http import require_POST
# PERBAIKAN: Tambahkan Cuti ke dalam import
from apps.hrd.models import Karyawan, JatahCuti, DetailJatahCuti, CutiBersama, Cuti
from apps.hrd.utils.jatah_cuti import (
    get_jatah_cuti_data, 
    update_manual_jatah_cuti, 
    get_expired_cuti_notifications,
    hitung_jatah_cuti
)
from django.db.models import Q
import calendar
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from datetime import datetime, timedelta
from django.core.paginator import Paginator
import json

@login_required
def laporan_jatah_cuti_view(request):
    """View untuk menampilkan laporan jatah cuti per karyawan per bulan."""
    if request.user.role != 'HRD':
        messages.error(request, "Anda tidak memiliki akses ke halaman ini.")
        return redirect('karyawan_dashboard')
    
    # Jika ini adalah ping untuk keep-alive session, kembalikan respons kosong
    if request.GET.get('ping') == 'true':
        return JsonResponse({'status': 'ok'})
    
    # Filter berdasarkan tahun
    tahun = request.GET.get('tahun', datetime.now().year)
    try:
        tahun = int(tahun)
    except (ValueError, TypeError):
        tahun = datetime.now().year
    
    # Filter berdasarkan nama karyawan (opsional)
    nama = request.GET.get('nama', '')
    
    # Ambil data jatah cuti menggunakan fungsi dari utils
    karyawan_id = None
    if nama:
        # Jika ada filter nama, cari karyawan yang sesuai
        karyawan_filtered = Karyawan.objects.filter(
            Q(user__role='HRD') | Q(user__role='Karyawan Tetap'),
            status_keaktifan='Aktif',
            nama__icontains=nama
        ).first()
        if karyawan_filtered:
            karyawan_id = karyawan_filtered.id
    
    # Ambil data laporan menggunakan fungsi dari utils
    laporan_data = get_jatah_cuti_data(tahun, karyawan_id)
    
    # Jika ada filter nama tapi tidak ditemukan karyawan, kosongkan data
    if nama and not karyawan_id:
        laporan_data = []
    
    # Ambil notifikasi cuti expired
    cuti_expired = get_expired_cuti_notifications(tahun)

    # jangan tampilkan jika user sudah menutupnya
    hide_key = f'hide_cuti_expired_{tahun}'
    if request.COOKIES.get(hide_key) == 'true':
        cuti_expired = []

    # Nama-nama bulan untuk header tabel
    nama_bulan = [calendar.month_name[i] for i in range(1, 13)]
    
    return render(request, 'hrd/laporan_jatah_cuti.html', {
        'laporan_data': laporan_data,
        'nama_bulan': nama_bulan,
        'tahun': tahun,
        'nama': nama,
        'cuti_expired': cuti_expired,
        'tahun_options': range(datetime.now().year - 2, datetime.now().year + 3)  # Opsi tahun: 2 tahun ke belakang, tahun ini, 2 tahun ke depan
    })

@login_required
def export_laporan_jatah_cuti_excel(request):
    """View untuk mengekspor laporan jatah cuti ke Excel."""
    if request.user.role != 'HRD':
        return HttpResponse("Forbidden", status=403)
    
    # Filter berdasarkan tahun
    tahun = request.GET.get('tahun', datetime.now().year)
    try:
        tahun = int(tahun)
    except (ValueError, TypeError):
        tahun = datetime.now().year
    
    # Filter berdasarkan nama karyawan (opsional)
    nama = request.GET.get('nama', '')
    
    # Ambil data jatah cuti menggunakan fungsi dari utils
    karyawan_id = None
    if nama:
        karyawan_filtered = Karyawan.objects.filter(
            Q(user__role='HRD') | Q(user__role='Karyawan Tetap'),
            status_keaktifan='Aktif',
            nama__icontains=nama
        ).first()
        if karyawan_filtered:
            karyawan_id = karyawan_filtered.id
    
    # Ambil data laporan menggunakan fungsi dari utils
    laporan_data = get_jatah_cuti_data(tahun, karyawan_id)
    
    # Jika ada filter nama tapi tidak ditemukan karyawan, kosongkan data
    if nama and not karyawan_id:
        laporan_data = []
    
    # Buat workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Laporan Jatah Cuti {tahun}"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    expired_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    used_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Header
    headers = ["No", "Nama Lengkap", "Januari", "Februari", "Maret", "April", "Mei", "Juni", 
              "Juli", "Agustus", "September", "Oktober", "November", "Desember", "Saldo Cuti"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row_num = 2
    
    for idx, data in enumerate(laporan_data, 1):
        karyawan = data['karyawan']
        bulan_data = data['bulan_data']
        sisa_cuti = data['sisa_cuti']
        
        # Tambahkan data ke Excel
        row = [idx, karyawan.nama]
        
        # Tambahkan data bulan
        for bulan_info in bulan_data:
            if bulan_info['dipakai']:
                # Ambil tanggal dari keterangan jika ada
                tanggal = bulan_info['keterangan'].split(': ')[-1] if ': ' in bulan_info['keterangan'] else ''
                row.append(tanggal)
            else:
                row.append('')
        
        # Tambahkan sisa cuti
        row.append(sisa_cuti)
        
        # Tulis ke Excel
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            
            # Styling untuk bulan
            if 2 < col_num < len(headers):
                cell.alignment = Alignment(horizontal='center')
                
                # Jika bulan dipakai, beri warna hijau
                if bulan_data[col_num-3]['dipakai']:
                    cell.fill = used_fill
                
                # Jika bulan expired, beri warna merah
                if bulan_data[col_num-3]['expired']:
                    cell.fill = expired_fill
        
        row_num += 1
    
    # Atur lebar kolom
    for col_num in range(1, len(headers) + 1):
        if col_num == 2:  # Kolom nama
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 30
        else:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
    
    # Freeze panes
    ws.freeze_panes = 'C2'
    
    # Buat response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=laporan_jatah_cuti_{tahun}.xlsx'
    wb.save(response)
    
    return response


@login_required
@require_POST
def update_jatah_cuti_ajax(request):
    """View untuk memperbarui detail jatah cuti via AJAX dengan logika sederhana."""
    if request.user.role != 'HRD':
        return JsonResponse({'status': 'error', 'message': 'Forbidden'}, status=403)
    
    try:
        # Handle multipart form data untuk file upload
        if request.content_type.startswith('multipart/form-data'):
            karyawan_id = request.POST.get('karyawan_id')
            tahun = request.POST.get('tahun')
            bulan = request.POST.get('bulan')
            dipakai = request.POST.get('dipakai') == 'true'
            keterangan = request.POST.get('keterangan', '')
            tanggal = request.POST.get('tanggal')  # Single date
            jenis_cuti = request.POST.get('jenis_cuti')
            file_persetujuan = request.FILES.get('file_persetujuan')
        else:
            data = json.loads(request.body)
            karyawan_id = data.get('karyawan_id')
            tahun = data.get('tahun')
            bulan = data.get('bulan')
            dipakai = data.get('dipakai', False)
            keterangan = data.get('keterangan', '')
            tanggal = data.get('tanggal')  # Single date
            jenis_cuti = data.get('jenis_cuti')
            file_persetujuan = None
        
        if not all([karyawan_id, tahun, bulan]):
            return JsonResponse({
                'success': False, 
                'message': 'Data tidak lengkap'
            }, status=400)
        
        # Panggil fungsi update dengan parameter baru
        result = update_manual_jatah_cuti(
            karyawan_id=int(karyawan_id),
            tahun=int(tahun),
            bulan=int(bulan),
            dipakai=dipakai,
            keterangan=keterangan,
            tanggal=tanggal,  # Single date instead of tanggal_mulai/tanggal_selesai
            jenis_cuti=jenis_cuti,
            file_persetujuan=file_persetujuan,
            user=request.user
        )
        
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Error: {str(e)}'
        }, status=500)

@login_required
def get_detail_jatah_cuti_ajax(request):
    """View untuk mendapatkan detail jatah cuti via AJAX dengan perbaikan prefill."""
    if request.user.role != 'HRD':
        return JsonResponse({'status': 'error', 'message': 'Forbidden'}, status=403)
    
    try:
        # Import lokal untuk model Cuti
        from apps.hrd.models import Cuti
        
        karyawan_id = request.GET.get('karyawan_id')
        tahun = request.GET.get('tahun')
        bulan = request.GET.get('bulan')
        
        if not all([karyawan_id, tahun, bulan]):
            return JsonResponse({'status': 'error', 'message': 'Data tidak lengkap'}, status=400)
        
        karyawan = get_object_or_404(Karyawan, id=karyawan_id)
        jatah_cuti = JatahCuti.objects.filter(karyawan=karyawan, tahun=int(tahun)).first()
        
        if not jatah_cuti:
            return JsonResponse({
                'status': 'success',
                'data': {
                    'dipakai': False, 'keterangan': '', 'tanggal': '',
                    'jenis_cuti': '', 'sisa_cuti': 0, 'expired': False
                }
            })
        
        # Ambil detail jatah cuti untuk bulan ini
        detail = DetailJatahCuti.objects.filter(
            jatah_cuti=jatah_cuti, tahun=int(tahun), bulan=int(bulan)
        ).first()
        
        response_data = {
            'dipakai': False, 'keterangan': '', 'tanggal': '',
            'jenis_cuti': '', 'sisa_cuti': jatah_cuti.sisa_cuti, 'expired': False
        }
        
        if detail and detail.dipakai:
            response_data.update({
                'dipakai': detail.dipakai,
                'keterangan': detail.keterangan,
            })
            
            # Untuk implementasi sederhana, gunakan tanggal_terpakai dari DetailJatahCuti
            if detail.tanggal_terpakai:
                response_data['tanggal'] = detail.tanggal_terpakai.strftime('%Y-%m-%d')
            
            # Cari data cuti terkait untuk mendapatkan jenis_cuti
            cuti_terkait = None
            
            # Metode 1: Cari berdasarkan tanggal_terpakai
            if detail.tanggal_terpakai:
                cuti_terkait = Cuti.objects.filter(
                    id_karyawan=karyawan,
                    tanggal_mulai=detail.tanggal_terpakai,
                    tanggal_selesai=detail.tanggal_terpakai,
                    status='disetujui'
                ).first()
            
            # Metode 2: Cari berdasarkan overlap dengan bulan ini
            if not cuti_terkait:
                from datetime import date
                first_day = date(int(tahun), int(bulan), 1)
                last_day = date(int(tahun), int(bulan), calendar.monthrange(int(tahun), int(bulan))[1])
                
                cuti_terkait = Cuti.objects.filter(
                    id_karyawan=karyawan,
                    status='disetujui'
                ).filter(
                    Q(tanggal_mulai__lte=last_day) & Q(tanggal_selesai__gte=first_day)
                ).order_by('-tanggal_mulai').first()
            
            if cuti_terkait:
                response_data.update({
                    'jenis_cuti': cuti_terkait.jenis_cuti,
                })
                
                # Untuk implementasi sederhana, gunakan tanggal_mulai sebagai tanggal tunggal
                if not response_data['tanggal']:
                    response_data['tanggal'] = cuti_terkait.tanggal_mulai.strftime('%Y-%m-%d')
        
        # Cek expired
        current_date = datetime.now().date()
        if int(tahun) < current_date.year - 1:
            response_data['expired'] = True
        elif int(tahun) == current_date.year - 1 and int(bulan) < current_date.month:
            response_data['expired'] = True
        
        return JsonResponse({'status': 'success', 'data': response_data})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def fix_jatah_cuti_slots(request):
    """View untuk memperbaiki slot jatah cuti yang salah posisi.
    
    Endpoint ini akan mencari semua DetailJatahCuti yang:
    - dipakai=True (sudah terisi cuti)
    - tersedia=False (bulan tidak tersedia berdasarkan kontrak)
    
    Lalu memindahkan data cuti tersebut ke slot yang benar (bulan pertama yang tersedia).
    """
    if request.user.role != 'HRD':
        messages.error(request, "Anda tidak memiliki akses ke halaman ini.")
        return redirect('karyawan_dashboard')
    
    from apps.hrd.utils.jatah_cuti import tentukan_bulan_tersedia_berdasarkan_kontrak
    
    hasil_perbaikan = []
    
    # Cari semua detail yang dipakai tapi tersedia=False (data yang salah)
    detail_salah = DetailJatahCuti.objects.filter(
        dipakai=True,
        tersedia=False
    ).select_related('jatah_cuti__karyawan')
    
    for detail in detail_salah:
        karyawan = detail.jatah_cuti.karyawan
        tahun = detail.tahun
        jatah_cuti = detail.jatah_cuti
        
        # Simpan data cuti yang akan dipindahkan
        keterangan_lama = detail.keterangan
        tanggal_terpakai_lama = detail.tanggal_terpakai
        jumlah_hari_lama = detail.jumlah_hari
        bulan_lama = detail.bulan
        
        # Cari slot kosong yang tersedia (bulan pertama yang available)
        slot_tersedia = DetailJatahCuti.objects.filter(
            jatah_cuti=jatah_cuti,
            tahun=tahun,
            dipakai=False,
            tersedia=True
        ).order_by('bulan').first()
        
        if slot_tersedia:
            # Pindahkan data ke slot yang benar
            slot_tersedia.dipakai = True
            slot_tersedia.jumlah_hari = jumlah_hari_lama
            slot_tersedia.keterangan = keterangan_lama
            slot_tersedia.tanggal_terpakai = tanggal_terpakai_lama
            slot_tersedia.save()
            
            # Kosongkan slot yang salah
            detail.dipakai = False
            detail.jumlah_hari = 0
            detail.keterangan = ''
            detail.tanggal_terpakai = None
            detail.save()
            
            hasil_perbaikan.append({
                'karyawan': karyawan.nama,
                'tahun': tahun,
                'dari_bulan': calendar.month_name[bulan_lama],
                'ke_bulan': calendar.month_name[slot_tersedia.bulan],
                'keterangan': keterangan_lama,
                'status': 'success'
            })
        else:
            hasil_perbaikan.append({
                'karyawan': karyawan.nama,
                'tahun': tahun,
                'dari_bulan': calendar.month_name[bulan_lama],
                'ke_bulan': None,
                'keterangan': keterangan_lama,
                'status': 'failed - tidak ada slot tersedia'
            })
    
    success_count = len([h for h in hasil_perbaikan if h["status"] == "success"])
    failed_count = len([h for h in hasil_perbaikan if h["status"] != "success"])
    
    return JsonResponse({
        'message': f'Perbaikan selesai. {success_count} slot berhasil diperbaiki, {failed_count} gagal.',
        'total_ditemukan': len(hasil_perbaikan),
        'berhasil': success_count,
        'gagal': failed_count,
        'hasil': hasil_perbaikan
    })