import pandas as pd
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Q, F
from django.utils import timezone
from apps.authentication.decorators import role_required
from apps.absensi.models import AbsensiMagang
from apps.hrd.models import Karyawan, Izin
from apps.authentication.models import User
from datetime import datetime, timedelta, time, date
import calendar
import openpyxl
from apps.hrd.utils.jatah_cuti import is_holiday_or_weekend
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def calculate_work_duration(jam_masuk, jam_pulang):
    """Menghitung durasi kerja dalam jam dari jam masuk dan pulang"""
    if not jam_masuk or not jam_pulang:
        return None
    
    # Combine with a dummy date for calculation
    today = datetime.today().date()
    start = datetime.combine(today, jam_masuk)
    end = datetime.combine(today, jam_pulang)
    
    # Handle cases where end time is before start (shouldn't happen but just in case)
    if end < start:
        return None
    
    duration = (end - start).total_seconds() / 3600
    return round(duration, 1)


@login_required
@role_required(['HRD'])
def riwayat_absensi_fleksibel_hr(request):
    """Dashboard Riwayat Absensi Fleksibel untuk HR - Control Center Kehadiran"""
    
    # Get selected date from request, default to today
    selected_date_str = request.GET.get('tanggal', '')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = datetime.now().date()
    else:
        selected_date = datetime.now().date()
    
    # Use selected_date for dashboard stats
    today = selected_date
    
    # ============================================
    # DASHBOARD STATISTICS
    # ============================================
    
    # Total karyawan aktif (always current, not filtered by date)
    total_karyawan_aktif = Karyawan.objects.filter(status_keaktifan='Aktif').count()
    
    # Absensi untuk tanggal yang dipilih (semua record)
    absensi_hari_ini = AbsensiMagang.objects.filter(tanggal=selected_date).select_related('id_karyawan', 'id_karyawan__user')
    
    # Hanya yang benar-benar check-in (jam_masuk terisi). Abaikan placeholder dari cron reminder.
    absensi_checkin = absensi_hari_ini.filter(jam_masuk__isnull=False)
    
    sudah_absen_hari_ini = absensi_checkin.count()
    wfo_hari_ini = absensi_checkin.filter(keterangan='WFO').count()
    wfa_hari_ini = absensi_checkin.filter(keterangan='WFA').count()
    
    # Karyawan yang belum absen pulang (sudah masuk tapi belum pulang)
    belum_pulang = absensi_checkin.filter(jam_pulang__isnull=True)
    belum_pulang_count = belum_pulang.count()
    belum_pulang_list = list(belum_pulang.values_list('id_karyawan__nama', flat=True)[:5])
    
    # Karyawan aktif yang belum absen masuk sama sekali hari ini 
    # Exclude: yang sudah check-in ATAU yang sudah punya catatan HR (hr_keterangan)
    karyawan_sudah_absen_ids = absensi_checkin.values_list('id_karyawan_id', flat=True)
    karyawan_dengan_catatan_hr = AbsensiMagang.objects.filter(
        tanggal=selected_date,
        hr_keterangan__isnull=False
    ).values_list('id_karyawan_id', flat=True)
    
    belum_masuk = Karyawan.objects.filter(
        status_keaktifan='Aktif'
    ).exclude(id__in=karyawan_sudah_absen_ids).exclude(id__in=karyawan_dengan_catatan_hr)
    belum_masuk_count = belum_masuk.count()
    belum_masuk_list = list(belum_masuk.values_list('nama', flat=True)[:5])
    
    # Full list of employees without attendance (for Add Note feature)
    belum_masuk_all = belum_masuk.select_related('user').order_by('nama')
    
    # ============================================
    # WORK DURATION INSIGHTS
    # ============================================
    
    # Query absensi yang sudah pulang hari ini (untuk hitung durasi)
    absensi_selesai = absensi_hari_ini.filter(jam_masuk__isnull=False, jam_pulang__isnull=False)
    
    # Hitung rata-rata durasi kerja
    total_durasi = 0
    durasi_list = []
    durasi_kurang_8_jam = []
    potensi_lembur = []
    
    for absensi in absensi_selesai:
        durasi = calculate_work_duration(absensi.jam_masuk, absensi.jam_pulang)
        if durasi:
            durasi_list.append(durasi)
            total_durasi += durasi
            
            if durasi < 8.5:
                durasi_kurang_8_jam.append({
                    'nama': absensi.id_karyawan.nama,
                    'durasi': durasi
                })
            elif durasi >= 10:
                potensi_lembur.append({
                    'nama': absensi.id_karyawan.nama,
                    'durasi': durasi
                })
    
    # Check for employees still working (belum pulang) with 10+ hours
    for absensi in belum_pulang:
        if absensi.jam_masuk:
            jam_masuk_dt = datetime.combine(selected_date, absensi.jam_masuk)
            sekarang = datetime.now()
            durasi = (sekarang - jam_masuk_dt).total_seconds() / 3600
            if durasi >= 10:
                potensi_lembur.append({
                    'nama': absensi.id_karyawan.nama,
                    'durasi': round(durasi, 1),
                    'masih_kerja': True
                })
    
    rata_rata_durasi = round(total_durasi / len(durasi_list), 1) if durasi_list else 0
    
    # ============================================
    # FILTERS
    # ============================================
    
    nama = request.GET.get('nama', '')
    bulan = request.GET.get('bulan')
    tahun = request.GET.get('tahun')
    role = request.GET.get('role', '')
    keterangan = request.GET.get('keterangan', '')
    tanggal = request.GET.get('tanggal', '')
    tanggal_mulai = request.GET.get('tanggal_mulai')
    tanggal_selesai = request.GET.get('tanggal_selesai')
    
    # Buat query dasar: 
    # - Yang punya jam_masuk (check-in), ATAU
    # - Yang punya hr_keterangan (catatan HR untuk tidak masuk)
    absensi_query = AbsensiMagang.objects.filter(
        Q(jam_masuk__isnull=False) | Q(hr_keterangan__isnull=False)
    ).select_related('id_karyawan', 'id_karyawan__user')
    
    # Terapkan filter
    if nama:
        absensi_query = absensi_query.filter(id_karyawan__nama__icontains=nama)
    
    if bulan and bulan.isdigit():
        absensi_query = absensi_query.filter(tanggal__month=int(bulan))
    
    if tahun and tahun.isdigit():
        absensi_query = absensi_query.filter(tanggal__year=int(tahun))
    
    if role:
        absensi_query = absensi_query.filter(id_karyawan__user__role=role)
    
    if keterangan:
        absensi_query = absensi_query.filter(keterangan=keterangan)
    
    if tanggal:
        try:
            tanggal_parsed = datetime.strptime(tanggal, '%Y-%m-%d').date()
            absensi_query = absensi_query.filter(tanggal=tanggal_parsed)
        except ValueError:
            messages.error(request, 'Format tanggal tidak valid')
    
    if tanggal_mulai:
        try:
            tanggal_mulai_parsed = datetime.strptime(tanggal_mulai, '%Y-%m-%d').date()
            absensi_query = absensi_query.filter(tanggal__gte=tanggal_mulai_parsed)
        except ValueError:
            messages.error(request, 'Format tanggal mulai tidak valid')
    
    if tanggal_selesai:
        try:
            tanggal_selesai_parsed = datetime.strptime(tanggal_selesai, '%Y-%m-%d').date()
            absensi_query = absensi_query.filter(tanggal__lte=tanggal_selesai_parsed)
        except ValueError:
            messages.error(request, 'Format tanggal selesai tidak valid')
    
    # Sorting with secondary sort
    sort_by = request.GET.get('sort_by', '-tanggal')
    # Add secondary sort by tanggal and jam_masuk for consistency
    if sort_by in ['id_karyawan__nama', '-id_karyawan__nama']:
        absensi_query = absensi_query.order_by(sort_by, '-tanggal', '-jam_masuk')
    elif sort_by in ['id_karyawan__user__role', '-id_karyawan__user__role']:
        absensi_query = absensi_query.order_by(sort_by, 'id_karyawan__nama', '-tanggal')
    elif sort_by in ['keterangan', '-keterangan']:
        absensi_query = absensi_query.order_by(sort_by, '-tanggal', '-jam_masuk')
    elif sort_by in ['jam_masuk', '-jam_masuk', 'jam_pulang', '-jam_pulang']:
        absensi_query = absensi_query.order_by(sort_by, 'id_karyawan__nama')
    else:
        # Default: sort by tanggal with secondary sort
        absensi_query = absensi_query.order_by(sort_by, '-jam_masuk')
    
    # ============================================
    # LATE LABEL (IZIN TELAT) LOGIC
    # ============================================
    #
    # Aturan:
    # - Label "Telat" hanya muncul di rekap bila ada Izin Telat DISERTAI
    #   created_at >= 10:00 untuk karyawan & tanggal tersebut.
    # - Izin Telat yang diajukan sebelum 10:00 tidak men-trigger label.
    #
    # Catatan:
    # - Check-in TANPA Izin Telat tetap tidak diperbolehkan oleh logic di views_fleksibel.
    #
    # Build a map of (karyawan_id, date) -> has_late_permission
    absensi_list_for_keys = list(absensi_query.values('id_karyawan_id', 'tanggal', 'id_absensi'))
    telat_label_map = {}
    if absensi_list_for_keys:
        karyawan_ids = list(set([item['id_karyawan_id'] for item in absensi_list_for_keys]))
        tanggal_list = list(set([item['tanggal'] for item in absensi_list_for_keys]))
        
        izin_telat_qs = Izin.objects.filter(
            jenis_izin='telat',
            status='disetujui',
            id_karyawan_id__in=karyawan_ids,
            tanggal_izin__in=tanggal_list,
        ).only('id_karyawan_id', 'tanggal_izin', 'created_at')
        
        cutoff_time = time(10, 0)
        for izin in izin_telat_qs:
            key = (izin.id_karyawan_id, izin.tanggal_izin)
            # Jika ada SATU saja izin dengan created_at >= 10:00 (waktu lokal),
            # maka label Telat perlu ditampilkan.
            if izin.created_at:
                # Konversi created_at ke timezone lokal (Asia/Jakarta)
                created_at_local = timezone.localtime(izin.created_at)
                created_time = created_at_local.time()
                if created_time >= cutoff_time:
                    telat_label_map[key] = True
                else:
                    telat_label_map.setdefault(key, False)
            else:
                telat_label_map.setdefault(key, False)
    
    # Calculate duration for each attendance record & attach late-label flag
    absensi_with_duration = []
    for absensi in absensi_query:
        durasi = calculate_work_duration(absensi.jam_masuk, absensi.jam_pulang)
        absensi.durasi_kerja = durasi
        
        key = (absensi.id_karyawan_id, absensi.tanggal)
        absensi.is_telat_label = telat_label_map.get(key, False)
        
        absensi_with_duration.append(absensi)
    
    # Pagination
    paginator = Paginator(absensi_with_duration, 15)  # 15 items per halaman
    page = request.GET.get('page')
    absensi_list = paginator.get_page(page)
    
    # ============================================
    # FILTER OPTIONS
    # ============================================
    
    months = [
        (1, 'Januari'), (2, 'Februari'), (3, 'Maret'), (4, 'April'),
        (5, 'Mei'), (6, 'Juni'), (7, 'Juli'), (8, 'Agustus'),
        (9, 'September'), (10, 'Oktober'), (11, 'November'), (12, 'Desember')
    ]
    
    current_year = datetime.now().year
    years = range(current_year - 3, current_year + 2)
    
    # Role choices dari model User
    role_choices = User.ROLE_CHOICES
    
    # Keterangan choices
    keterangan_choices = [
        ('WFO', 'WFO'),
        ('WFA', 'WFA'),
        ('Izin Telat', 'Izin Telat'),
        ('Izin Sakit', 'Izin Sakit')
    ]
    
    # ============================================
    # PIVOT TABLE (untuk rekap bulanan)
    # ============================================
    
    pivot_headers = []
    pivot_rows = []
    
    if bulan and bulan.isdigit() and tahun and tahun.isdigit():
        if absensi_query.exists():
            pivot_data = list(absensi_query.values('id_karyawan__nama', 'keterangan'))
            df = pd.DataFrame(pivot_data)
            
            if not df.empty and 'id_karyawan__nama' in df.columns and 'keterangan' in df.columns:
                pivot_table = pd.pivot_table(
                    df,
                    index='id_karyawan__nama',
                    columns='keterangan',
                    aggfunc='size',
                    fill_value=0
                ).reset_index()
                
                all_keterangan = ['WFO', 'WFA']
                for ket in all_keterangan:
                    if ket not in pivot_table.columns:
                        pivot_table[ket] = 0
                
                numeric_columns = [col for col in pivot_table.columns if col != 'id_karyawan__nama']
                pivot_table['Total'] = pivot_table[numeric_columns].sum(axis=1)
                
                column_order = ['id_karyawan__nama'] + all_keterangan + ['Total']
                pivot_table = pivot_table.reindex(columns=column_order, fill_value=0)
                
                pivot_headers = ['Nama Karyawan'] + all_keterangan + ['Total']
                pivot_rows = [list(row) for _, row in pivot_table.iterrows()]
    
    # ============================================
    # REKAP HARI KERJA (mirip Jatah Cuti, kolom per hari kerja)
    # ============================================
    
    rekap_hari_kerja_headers = []
    rekap_hari_kerja_rows = []
    hari_kerja_list = []
    
    if bulan and bulan.isdigit() and tahun and tahun.isdigit():
        bulan_int = int(bulan)
        tahun_int = int(tahun)
        
        # Daftar hari kerja dalam bulan (tanggal yang bukan libur/weekend)
        first_day = date(tahun_int, bulan_int, 1)
        _, last_day_num = calendar.monthrange(tahun_int, bulan_int)
        last_day = date(tahun_int, bulan_int, last_day_num)
        
        current = first_day
        while current <= last_day:
            if not is_holiday_or_weekend(current):
                hari_kerja_list.append({
                    'tanggal': current,
                    'tanggal_str': current.strftime('%Y-%m-%d'),
                    'label': str(current.day),
                })
            current += timedelta(days=1)
        
        # Query absensi untuk bulan ini (sama filter dengan pivot)
        rekap_absensi_query = AbsensiMagang.objects.filter(
            Q(jam_masuk__isnull=False) | Q(hr_keterangan__isnull=False),
            tanggal__month=bulan_int,
            tanggal__year=tahun_int,
        ).select_related('id_karyawan', 'id_karyawan__user')
        
        if nama:
            rekap_absensi_query = rekap_absensi_query.filter(id_karyawan__nama__icontains=nama)
        if role:
            rekap_absensi_query = rekap_absensi_query.filter(id_karyawan__user__role=role)
        
        # Karyawan unik yang punya data di bulan ini
        karyawan_ids = list(rekap_absensi_query.values_list('id_karyawan_id', flat=True).distinct())
        karyawan_rekap = Karyawan.objects.filter(id__in=karyawan_ids).order_by('nama') if karyawan_ids else []
        
        # Map absensi per (karyawan_id, tanggal)
        absensi_by_key = {}
        for a in rekap_absensi_query:
            key = (a.id_karyawan_id, a.tanggal)
            absensi_by_key[key] = a
        
        # Izin telat (untuk label Telat) - cutoff 10:00
        izin_telat_map = {}
        if karyawan_ids and hari_kerja_list:
            tanggal_set = {h['tanggal'] for h in hari_kerja_list}
            for izin in Izin.objects.filter(
                jenis_izin='telat',
                status='disetujui',
                id_karyawan_id__in=karyawan_ids,
                tanggal_izin__in=tanggal_set,
            ).only('id_karyawan_id', 'tanggal_izin', 'created_at'):
                if izin.created_at:
                    created_local = timezone.localtime(izin.created_at)
                    if created_local.time() >= time(10, 0):
                        izin_telat_map[(izin.id_karyawan_id, izin.tanggal_izin)] = True
        
        # Izin pulang awal
        izin_pulang_awal_map = set()
        for izin in Izin.objects.filter(
            jenis_izin='pulang_awal',
            status='disetujui',
            id_karyawan_id__in=karyawan_ids,
            tanggal_izin__month=bulan_int,
            tanggal_izin__year=tahun_int,
        ).values_list('id_karyawan_id', 'tanggal_izin'):
            izin_pulang_awal_map.add((izin[0], izin[1]))
        
        # Build rows
        for karyawan in karyawan_rekap:
            total_hadir = 0
            cells = []
            for hari in hari_kerja_list:
                tgl = hari['tanggal']
                key = (karyawan.id, tgl)
                absensi = absensi_by_key.get(key)
                
                if not absensi:
                    cells.append({
                        'label': '-',
                        'punya_detail': False,
                        'tanggal_str': hari['tanggal_str'],
                        'badge_class': '',
                    })
                    continue
                
                # Tentukan label
                labels = []
                badge_class = 'badge-secondary'
                
                if absensi.hr_keterangan and not absensi.jam_masuk:
                    labels.append('Catatan HR')
                    badge_class = 'badge-secondary'
                elif absensi.keterangan:
                    labels.append(absensi.keterangan)
                    if absensi.keterangan == 'WFO':
                        badge_class = 'badge-primary'
                    elif absensi.keterangan == 'WFA':
                        badge_class = 'badge-info'
                    elif absensi.keterangan in ('Izin Telat', 'Izin Sakit'):
                        badge_class = 'badge-warning' if absensi.keterangan == 'Izin Telat' else 'badge-danger'
                
                if absensi.jam_masuk and not absensi.jam_pulang and absensi.keterangan != 'Tidak Masuk':
                    labels.append('Belum Pulang')
                    badge_class = 'badge-warning'
                
                if izin_telat_map.get(key):
                    labels.append('Telat')
                    if 'badge-warning' not in badge_class:
                        badge_class = 'badge-warning'
                
                if key in izin_pulang_awal_map:
                    labels.append('Pulang Awal')
                    badge_class = 'badge-warning'
                
                label_display = ', '.join(labels) if labels else '-'
                
                # punya_detail: tampilkan popup untuk cell yang punya data (telat, WFA, tidak lengkap, pulang awal, atau absensi)
                punya_detail = True
                
                if absensi.jam_masuk:
                    total_hadir += 1
                
                cells.append({
                    'label': label_display,
                    'punya_detail': punya_detail,
                    'tanggal_str': hari['tanggal_str'],
                    'badge_class': badge_class,
                })
            
            rekap_hari_kerja_rows.append({
                'karyawan': karyawan,
                'karyawan_id': karyawan.id,
                'cells': cells,
                'total': total_hadir,
            })
        
        rekap_hari_kerja_headers = ['No', 'Nama Karyawan'] + [h['label'] for h in hari_kerja_list] + ['Total']
    
    # ============================================
    # CONTEXT
    # ============================================
    
    context = {
        # Dashboard stats
        'today': selected_date,
        'total_karyawan_aktif': total_karyawan_aktif,
        'sudah_absen_hari_ini': sudah_absen_hari_ini,
        'wfo_hari_ini': wfo_hari_ini,
        'wfa_hari_ini': wfa_hari_ini,
        'belum_masuk_count': belum_masuk_count,
        'belum_masuk_list': belum_masuk_list,
        'belum_masuk_all': belum_masuk_all,  # Full queryset for Add Note feature
        'belum_pulang_count': belum_pulang_count,
        'belum_pulang_list': belum_pulang_list,
        'rata_rata_durasi': rata_rata_durasi,
        'durasi_kurang_8_jam': durasi_kurang_8_jam,
        'potensi_lembur': potensi_lembur,
        
        # Table data
        'absensi_list': absensi_list,
        
        # Filter values
        'nama_filter': nama,
        'selected_month': int(bulan) if bulan and bulan.isdigit() else '',
        'selected_year': int(tahun) if tahun and tahun.isdigit() else '',
        'selected_role': role,
        'selected_keterangan': keterangan,
        'tanggal_filter': tanggal,
        'tanggal_mulai': tanggal_mulai,
        'tanggal_selesai': tanggal_selesai,
        'sort_by': sort_by,
        
        # Filter options
        'months': months,
        'years': years,
        'role_choices': role_choices,
        'keterangan_choices': keterangan_choices,
        
        # Pivot table
        'pivot_headers': pivot_headers,
        'pivot_rows': pivot_rows,
        
        # Rekap hari kerja
        'rekap_hari_kerja_headers': rekap_hari_kerja_headers,
        'rekap_hari_kerja_rows': rekap_hari_kerja_rows,
        'hari_kerja_list': hari_kerja_list,
        
        # Page title
        'title': 'Riwayat Absensi',
        'selected_date': selected_date
    }
    
    return render(request, 'absensi/riwayat_absensi_fleksibel_hr.html', context)


@login_required
@role_required(['HRD'])
def export_absensi_fleksibel_excel(request):
    """Ekspor data absensi fleksibel ke Excel dengan Role dan Durasi"""
    nama = request.GET.get('nama', '')
    bulan = request.GET.get('bulan')
    tahun = request.GET.get('tahun')
    role = request.GET.get('role', '')
    keterangan = request.GET.get('keterangan', '')
    tanggal_mulai = request.GET.get('tanggal_mulai')
    tanggal_selesai = request.GET.get('tanggal_selesai')
    
    # Buat query dasar
    absensi_query = AbsensiMagang.objects.all().select_related('id_karyawan', 'id_karyawan__user')
    
    # Terapkan filter
    if nama:
        absensi_query = absensi_query.filter(id_karyawan__nama__icontains=nama)
    
    if bulan and bulan.isdigit():
        absensi_query = absensi_query.filter(tanggal__month=int(bulan))
    
    if tahun and tahun.isdigit():
        absensi_query = absensi_query.filter(tanggal__year=int(tahun))
    
    if role:
        absensi_query = absensi_query.filter(id_karyawan__user__role=role)
    
    if keterangan:
        absensi_query = absensi_query.filter(keterangan=keterangan)
    
    if tanggal_mulai:
        try:
            tanggal_mulai = datetime.strptime(tanggal_mulai, '%Y-%m-%d').date()
            absensi_query = absensi_query.filter(tanggal__gte=tanggal_mulai)
        except ValueError:
            pass
    
    if tanggal_selesai:
        try:
            tanggal_selesai = datetime.strptime(tanggal_selesai, '%Y-%m-%d').date()
            absensi_query = absensi_query.filter(tanggal__lte=tanggal_selesai)
        except ValueError:
            pass
    
    absensi_query = absensi_query.order_by('-tanggal')
    
    # Label Telat (1/0): sama dengan logic di riwayat_absensi_fleksibel_hr
    absensi_keys = list(absensi_query.values_list('id_karyawan_id', 'tanggal').distinct())
    telat_label_map = {}
    if absensi_keys:
        karyawan_ids = [k for k, _ in absensi_keys]
        tanggal_list = [t for _, t in absensi_keys]
        izin_telat_qs = Izin.objects.filter(
            jenis_izin='telat',
            status='disetujui',
            id_karyawan_id__in=karyawan_ids,
            tanggal_izin__in=tanggal_list,
        ).only('id_karyawan_id', 'tanggal_izin', 'created_at')
        cutoff_time = time(10, 0)
        for izin in izin_telat_qs:
            key = (izin.id_karyawan_id, izin.tanggal_izin)
            if izin.created_at:
                created_at_local = timezone.localtime(izin.created_at)
                created_time = created_at_local.time()
                if created_time >= cutoff_time:
                    telat_label_map[key] = True
                else:
                    telat_label_map.setdefault(key, False)
            else:
                telat_label_map.setdefault(key, False)
    
    # Buat workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Riwayat Absensi"
    
    # Style untuk header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header
    headers = [
        "Nama Karyawan", "Role", "Tanggal", "Jam Masuk", "Jam Pulang",
        "Durasi (Jam)", "Keterangan", "Telat", "Lokasi Masuk", "Lokasi Pulang", "Catatan HR",
        "Auto CO", "Alasan Lupa CO", "Jam Pulang Kira"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data
    for row_num, absensi in enumerate(absensi_query, 2):
        durasi = calculate_work_duration(absensi.jam_masuk, absensi.jam_pulang)
        role_name = absensi.id_karyawan.user.role if absensi.id_karyawan.user else "-"
        is_telat = telat_label_map.get((absensi.id_karyawan_id, absensi.tanggal), False)
        
        row_data = [
            absensi.id_karyawan.nama,
            role_name,
            absensi.tanggal.strftime("%Y-%m-%d"),
            absensi.jam_masuk.strftime("%H:%M:%S") if absensi.jam_masuk else "-",
            absensi.jam_pulang.strftime("%H:%M:%S") if absensi.jam_pulang else "-",
            durasi if durasi else "-",
            absensi.keterangan or "-",
            1 if is_telat else 0,
            absensi.alamat_masuk or "-",
            absensi.alamat_pulang or "-",
            absensi.hr_keterangan or "-",
            "Ya" if getattr(absensi, 'co_auto_generated', False) else "-",
            getattr(absensi, 'alasan_lupa_co', None) or "-",
            absensi.jam_pulang_kira.strftime("%H:%M") if getattr(absensi, 'jam_pulang_kira', None) else "-"
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col in [3, 4, 5, 6, 7, 8]:  # Center align date, time, duration, keterangan, telat
                cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Buat respons download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"riwayat_absensi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response


@login_required
@role_required(['HRD'])
def export_rekap_absensi_fleksibel_excel(request):
    """Ekspor rekap pivot absensi fleksibel ke Excel"""
    bulan = request.GET.get('bulan')
    tahun = request.GET.get('tahun')
    nama = request.GET.get('nama', '')
    role = request.GET.get('role', '')
    
    if not bulan or not tahun:
        messages.error(request, "Bulan dan tahun harus dipilih untuk ekspor rekap.")
        return redirect("riwayat_absensi_fleksibel_hr")
    
    # Buat query dasar
    absensi_query = AbsensiMagang.objects.filter(
        tanggal__month=int(bulan),
        tanggal__year=int(tahun)
    ).select_related('id_karyawan', 'id_karyawan__user')
    
    if nama:
        absensi_query = absensi_query.filter(id_karyawan__nama__icontains=nama)
    
    if role:
        absensi_query = absensi_query.filter(id_karyawan__user__role=role)
    
    # Buat workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Absensi"
    
    # Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    if absensi_query.exists():
        pivot_data = list(absensi_query.values('id_karyawan__nama', 'id_karyawan__user__role', 'keterangan'))
        df = pd.DataFrame(pivot_data)
        
        if not df.empty:
            pivot_table = pd.pivot_table(
                df,
                index=['id_karyawan__nama', 'id_karyawan__user__role'],
                columns='keterangan',
                aggfunc='size',
                fill_value=0
            ).reset_index()
            
            # Ensure WFO and WFA columns exist
            for ket in ['WFO', 'WFA']:
                if ket not in pivot_table.columns:
                    pivot_table[ket] = 0
            
            numeric_columns = [col for col in pivot_table.columns if col not in ['id_karyawan__nama', 'id_karyawan__user__role']]
            pivot_table['Total'] = pivot_table[numeric_columns].sum(axis=1)
            
            # Rename columns for clarity
            pivot_table = pivot_table.rename(columns={
                'id_karyawan__nama': 'Nama Karyawan',
                'id_karyawan__user__role': 'Role'
            })
            
            # Header
            headers = list(pivot_table.columns)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data
            for row_num, (_, row) in enumerate(pivot_table.iterrows(), 2):
                for col, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col, value=value)
    else:
        headers = ['Nama Karyawan', 'Role', 'WFO', 'WFA', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Buat respons download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    bulan_nama = dict([
        (1, 'Januari'), (2, 'Februari'), (3, 'Maret'), (4, 'April'),
        (5, 'Mei'), (6, 'Juni'), (7, 'Juli'), (8, 'Agustus'),
        (9, 'September'), (10, 'Oktober'), (11, 'November'), (12, 'Desember')
    ]).get(int(bulan), str(bulan))
    filename = f"rekap_absensi_{bulan_nama}_{str(tahun)}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response


@login_required
@role_required(['HRD'])
def get_detail_absensi_hari_ajax(request):
    """Detail absensi untuk satu karyawan pada satu tanggal (untuk modal Rekap Hari Kerja)."""
    if request.method != 'GET':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

    karyawan_id = request.GET.get('karyawan_id')
    tanggal_str = request.GET.get('tanggal')

    if not karyawan_id or not tanggal_str:
        return JsonResponse({'status': 'error', 'message': 'Data tidak lengkap'}, status=400)

    try:
        tanggal = datetime.strptime(tanggal_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Format tanggal tidak valid'}, status=400)

    karyawan = Karyawan.objects.filter(id=karyawan_id).first()
    if not karyawan:
        return JsonResponse({'status': 'error', 'message': 'Karyawan tidak ditemukan'}, status=404)

    absensi = AbsensiMagang.objects.filter(
        id_karyawan=karyawan,
        tanggal=tanggal
    ).first()

    durasi = None
    if absensi and absensi.jam_masuk and absensi.jam_pulang:
        durasi = calculate_work_duration(absensi.jam_masuk, absensi.jam_pulang)

    # Cek izin telat (created_at >= 10:00)
    is_telat = False
    izin_telat = Izin.objects.filter(
        jenis_izin='telat',
        status='disetujui',
        id_karyawan=karyawan,
        tanggal_izin=tanggal
    ).first()
    if izin_telat and izin_telat.created_at:
        created_local = timezone.localtime(izin_telat.created_at)
        if created_local.time() >= time(10, 0):
            is_telat = True

    # Cek izin pulang awal
    izin_pulang_awal = Izin.objects.filter(
        jenis_izin='pulang_awal',
        status='disetujui',
        id_karyawan=karyawan,
        tanggal_izin=tanggal
    ).first()

    data = {
        'nama_karyawan': karyawan.nama,
        'tanggal': tanggal_str,
        'tanggal_display': tanggal.strftime('%d %b %Y'),
        'jam_masuk': absensi.jam_masuk.strftime('%H:%M') if absensi and absensi.jam_masuk else '-',
        'jam_pulang': absensi.jam_pulang.strftime('%H:%M') if absensi and absensi.jam_pulang else '-',
        'durasi': f'{durasi} jam' if durasi else '-',
        'keterangan': absensi.keterangan or '-' if absensi else '-',
        'hr_keterangan': absensi.hr_keterangan or '-' if absensi else '-',
        'is_telat': is_telat,
        'izin_telat_alasan': izin_telat.alasan if izin_telat else None,
        'izin_pulang_awal': bool(izin_pulang_awal),
        'izin_pulang_awal_alasan': izin_pulang_awal.alasan if izin_pulang_awal else None,
        'belum_pulang': bool(absensi and absensi.jam_masuk and not absensi.jam_pulang),
        'co_auto_generated': getattr(absensi, 'co_auto_generated', False) if absensi else False,
        'alasan_lupa_co': getattr(absensi, 'alasan_lupa_co', None) or '-' if absensi else '-',
        'aktivitas_wfa': absensi.aktivitas_wfa or '-' if absensi else '-',
    }

    return JsonResponse({'status': 'success', 'data': data})


@login_required
@role_required(['HRD'])
def save_hr_attendance_note(request):
    """
    Save HR's free-text note for employee with no attendance activity.
    Creates or updates AbsensiMagang record with hr_keterangan field.
    Accessed via AJAX from /absensi/fleksibel-hr/ page.
    """
    if request.method == 'POST':
        karyawan_id = request.POST.get('karyawan_id')
        tanggal_str = request.POST.get('tanggal')
        keterangan = request.POST.get('hr_keterangan', '').strip()
        
        # Validation
        if not karyawan_id:
            return JsonResponse({
                'status': 'error',
                'message': 'ID karyawan tidak valid'
            }, status=400)
        
        if not tanggal_str:
            return JsonResponse({
                'status': 'error',
                'message': 'Tanggal tidak valid'
            }, status=400)
        
        if not keterangan:
            return JsonResponse({
                'status': 'error',
                'message': 'Keterangan tidak boleh kosong'
            }, status=400)
        
        try:
            karyawan = get_object_or_404(Karyawan, id=karyawan_id)
            tanggal = datetime.strptime(tanggal_str, '%Y-%m-%d').date()
        except (Karyawan.DoesNotExist, ValueError) as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Data tidak valid: {str(e)}'
            }, status=400)
        
        # Get or create attendance record for this employee on this date
        absensi, created = AbsensiMagang.objects.get_or_create(
            id_karyawan=karyawan,
            tanggal=tanggal,
            defaults={
                'hr_keterangan': keterangan,
                'keterangan': 'Tidak Masuk'  # Only for new records without check-in
            }
        )
        
        if not created:
            # Update existing record - ONLY update hr_keterangan, preserve existing keterangan
            absensi.hr_keterangan = keterangan
            # Do NOT overwrite keterangan if record already exists (already has WFO/WFA/etc)
            absensi.save()
        
        return JsonResponse({
            'status': 'success',
            'message': f'Catatan untuk {karyawan.nama} berhasil disimpan',
            'created': created,
            'keterangan': keterangan
        })
    
    return JsonResponse({
        'status': 'error',
        'message': 'Method not allowed'
    }, status=405)